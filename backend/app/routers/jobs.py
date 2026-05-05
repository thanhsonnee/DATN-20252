from __future__ import annotations

import io
import multiprocessing

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session, joinedload

from app.core.deps import get_current_user, require_algo_tester
from app.db.models import Job, JobStatus, User
from app.db.session import SessionLocal, get_db
from app.models.schemas import JobCreate, JobOut
from app.services.solver_service import run_solver_job

router = APIRouter(prefix="/jobs", tags=["jobs"])


def _solver_process_main(job_id: int) -> None:
    """Entry point for solver subprocess — creates its own DB session."""
    db = SessionLocal()
    try:
        run_solver_job(job_id, db)
    finally:
        db.close()


def _run_in_background(job_id: int) -> None:
    """Spawn a separate process so the CPU-bound solver doesn't block the GIL
    and freeze all other HTTP requests while the job is running."""
    p = multiprocessing.Process(target=_solver_process_main, args=(job_id,), daemon=True)
    p.start()


@router.post("/", response_model=JobOut, status_code=status.HTTP_201_CREATED)
def create_job(
    body: JobCreate,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current: User = Depends(require_algo_tester),
):
    job = Job(
        instance_name=body.instance_name,
        method=body.method,
        time_limit_sec=body.time_limit_sec,
        seed=body.seed,
        owner_id=current.id,
        status=JobStatus.pending,
    )
    db.add(job)
    db.commit()
    db.refresh(job)
    background_tasks.add_task(_run_in_background, job.id)
    return job


@router.get("/", response_model=list[JobOut])
def list_jobs(db: Session = Depends(get_db), current: User = Depends(require_algo_tester)):
    return (
        db.query(Job)
        .options(joinedload(Job.solution))
        .order_by(Job.created_at.desc())
        .all()
    )


@router.get("/export")
def export_jobs_excel(
    db: Session = Depends(get_db),
    current: User = Depends(require_algo_tester),
):
    from app.db.models import Solution, UserRole

    q = db.query(Job).filter(Job.status == JobStatus.done).options(joinedload(Job.solution))
    if current.role != UserRole.admin:
        q = q.filter(Job.owner_id == current.id)
    jobs = q.order_by(Job.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"

    headers = [
        "ID", "Instance", "Dataset", "Method", "Seed", "Time Limit (s)",
        "Started", "Finished", "Duration (s)",
        "Vehicles (NV)", "Total Distance", "Total Cost",
        "Iterations", "Runtime (s)", "Iter/sec",
    ]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for job in jobs:
        sol: Solution | None = job.solution
        duration = None
        if job.started_at and job.finished_at:
            duration = round((job.finished_at - job.started_at).total_seconds(), 2)
        iter_per_sec = None
        if sol and sol.iterations and sol.elapsed_sec and sol.elapsed_sec > 0:
            iter_per_sec = round(sol.iterations / sol.elapsed_sec, 1)

        is_lilim = sol and sol.dataset_type == "lilim"
        # Li&Lim: value in Total Distance, "x" in Total Cost
        # Sartori/others: "x" in Total Distance, value in Total Cost
        if sol:
            if is_lilim:
                td_val = round(sol.total_distance, 4)
                tc_val = "x"
            else:
                td_val = "x"
                tc_val = round(sol.total_cost, 4) if sol.total_cost is not None else round(sol.total_distance, 4)
        else:
            td_val = None
            tc_val = None

        ws.append([
            job.id,
            job.instance_name,
            sol.dataset_type if sol and sol.dataset_type else "",
            job.method,
            job.seed,
            job.time_limit_sec,
            job.started_at.strftime("%Y-%m-%d %H:%M:%S") if job.started_at else "",
            job.finished_at.strftime("%Y-%m-%d %H:%M:%S") if job.finished_at else "",
            duration,
            sol.num_vehicles if sol else None,
            td_val,
            tc_val,
            sol.iterations if sol else None,
            round(sol.elapsed_sec, 2) if sol and sol.elapsed_sec is not None else None,
            iter_per_sec,
        ])

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=jobs_export.xlsx"},
    )


@router.get("/{job_id}", response_model=JobOut)
def get_job(job_id: int, db: Session = Depends(get_db), current: User = Depends(require_algo_tester)):
    job = db.query(Job).options(joinedload(Job.solution)).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    return job


@router.post("/{job_id}/cancel", response_model=JobOut)
def cancel_job(
    job_id: int,
    db: Session = Depends(get_db),
    current: User = Depends(require_algo_tester),
):
    job = db.query(Job).options(joinedload(Job.solution)).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.status not in (JobStatus.pending, JobStatus.running):
        raise HTTPException(
            status_code=400,
            detail=f"Không thể huỷ job ở trạng thái '{job.status.value}'"
        )
    job.status = JobStatus.failed
    job.error_msg = "Cancelled by user"
    db.commit()
    db.refresh(job)
    return job


@router.delete("/{job_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_job(job_id: int, db: Session = Depends(get_db), current: User = Depends(require_algo_tester)):
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    if job.status == JobStatus.running:
        raise HTTPException(status_code=400, detail="Cannot delete a running job")
    db.delete(job)
    db.commit()
